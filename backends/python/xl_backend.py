#!/usr/bin/env python3
"""Python/openpyxl backend for @yome/xl on headless hosts.

The CLI runner talks to this script over a tiny JSON stdin/stdout protocol.
Workbook bytes remain the source of truth; this backend keeps only lightweight
"active workbook / active sheet" state under ~/.yome/state.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import time
from pathlib import Path
from typing import Any, Iterable

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter, range_boundaries
except Exception as exc:  # pragma: no cover - exercised through --probe
    openpyxl = None  # type: ignore[assignment]
    _IMPORT_ERROR = exc
else:
    _IMPORT_ERROR = None


BACKEND_ID = "python-openpyxl"
STATE_ROOT = Path(os.environ.get("YOME_STATE_HOME") or (Path.home() / ".yome" / "state"))
STATE_FILE = STATE_ROOT / "@yome" / "xl" / "linux-session.json"
SUPPORTED_ACTIONS = [
    "info",
    "open",
    "new",
    "save",
    "close",
    "books",
    "sheets",
    "sheet",
    "sheet.add",
    "sheet.rename",
    "sheet.delete",
    "used",
    "get",
    "range",
    "find",
    "set",
    "fill",
    "clear",
    "fmt",
    "width",
    "row.add",
    "row.delete",
    "col.add",
    "col.delete",
    "merge",
    "unmerge",
    "export",
]

COLOR_NAMES = {
    "black": "FF000000",
    "white": "FFFFFFFF",
    "red": "FFFF0000",
    "green": "FF008000",
    "blue": "FF0000FF",
    "yellow": "FFFFFF00",
    "gray": "FF808080",
    "grey": "FF808080",
    "orange": "FFFFA500",
    "purple": "FF800080",
}


def main() -> int:
    parser = argparse.ArgumentParser(add_help=False)
    parser.add_argument("--probe", action="store_true")
    parser.add_argument("--dispatch", action="store_true")
    ns = parser.parse_args()

    if ns.probe:
        return emit_probe()
    if ns.dispatch:
        if _IMPORT_ERROR is not None:
            return emit_result(False, stderr=dependency_error(), exit_code=127)
        try:
            payload = json.load(sys.stdin)
            result = dispatch(payload)
            return emit_result(**result)
        except Exception as exc:
            return emit_result(False, stderr=f"xl python backend: {exc}", exit_code=1)

    return emit_result(False, stderr="usage: xl_backend.py --probe | --dispatch", exit_code=2)


def emit_probe() -> int:
    if _IMPORT_ERROR is not None:
        return emit_json({"ok": False, "backend": BACKEND_ID, "stderr": dependency_error()})
    return emit_json(
        {
            "ok": True,
            "backend": BACKEND_ID,
            "engine": "openpyxl",
            "openpyxl": getattr(openpyxl, "__version__", "unknown"),
            "supports": SUPPORTED_ACTIONS,
        }
    )


def emit_result(ok: bool, stdout: str = "", stderr: str = "", exit_code: int | None = None, **extra: Any) -> int:
    code = exit_code if exit_code is not None else (0 if ok else 1)
    return emit_json({"ok": ok, "stdout": stdout, "stderr": stderr, "exitCode": code, **extra}, code)


def emit_json(obj: dict[str, Any], exit_code: int = 0) -> int:
    sys.stdout.write(json.dumps(obj, ensure_ascii=False))
    sys.stdout.write("\n")
    return exit_code


def dependency_error() -> str:
    return (
        "python-openpyxl backend requires openpyxl. Install it with "
        "`python3 -m pip install openpyxl`, or set YOME_PYTHON to a Python "
        f"environment that has openpyxl. Import error: {_IMPORT_ERROR}"
    )


def dispatch(req: dict[str, Any]) -> dict[str, Any]:
    action = str(req.get("action") or "")
    positionals = [str(v) for v in req.get("positionals") or []]
    flags = {str(k): v for k, v in (req.get("flags") or {}).items()}
    cwd = str(req.get("workingDirectory") or os.environ.get("YOME_WORKING_DIRECTORY") or os.getcwd())
    state = load_state()

    if action == "info":
        return with_workbook(state, flags, cwd, positionals, lambda wb, path, pos, fx: xl_info(wb, path, fx, state))
    if action == "open":
        return xl_open(positionals, flags, cwd, state)
    if action == "new":
        return xl_new(positionals, flags, cwd, state)
    if action == "save":
        return xl_save(positionals, flags, cwd, state)
    if action == "close":
        return xl_close(flags, state)
    if action == "books":
        return xl_books(state)
    if action == "sheets":
        return with_workbook(state, flags, cwd, positionals, lambda wb, path, pos, fx: xl_sheets(wb, path))
    if action == "sheet":
        return with_workbook(state, flags, cwd, positionals, lambda wb, path, pos, fx: xl_sheet(wb, path, pos, state))
    if action == "sheet.add":
        return with_workbook(state, flags, cwd, positionals, lambda wb, path, pos, fx: xl_sheet_add(wb, path, fx, state))
    if action == "sheet.rename":
        return with_workbook(state, flags, cwd, positionals, lambda wb, path, pos, fx: xl_sheet_rename(wb, path, pos, fx, state))
    if action == "sheet.delete":
        return with_workbook(state, flags, cwd, positionals, lambda wb, path, pos, fx: xl_sheet_delete(wb, path, pos, state))
    if action == "used":
        return with_workbook(state, flags, cwd, positionals, lambda wb, path, pos, fx: xl_used(wb, path, fx, state))
    if action == "get":
        return with_workbook(state, flags, cwd, positionals, lambda wb, path, pos, fx: xl_get(wb, path, pos, fx, state))
    if action == "range":
        return with_workbook(state, flags, cwd, positionals, lambda wb, path, pos, fx: xl_range(wb, path, pos, fx, state))
    if action == "find":
        return with_workbook(state, flags, cwd, positionals, lambda wb, path, pos, fx: xl_find(wb, path, pos, fx, state))
    if action == "set":
        return with_workbook(state, flags, cwd, positionals, lambda wb, path, pos, fx: xl_set(wb, path, pos, fx, state))
    if action == "fill":
        return with_workbook(state, flags, cwd, positionals, lambda wb, path, pos, fx: xl_fill(wb, path, pos, fx, state))
    if action == "clear":
        return with_workbook(state, flags, cwd, positionals, lambda wb, path, pos, fx: xl_clear(wb, path, pos, fx, state))
    if action == "fmt":
        return with_workbook(state, flags, cwd, positionals, lambda wb, path, pos, fx: xl_fmt(wb, path, pos, fx, state))
    if action == "width":
        return with_workbook(state, flags, cwd, positionals, lambda wb, path, pos, fx: xl_width(wb, path, pos, fx, state))
    if action == "row.add":
        return with_workbook(state, flags, cwd, positionals, lambda wb, path, pos, fx: xl_row_add(wb, path, pos, fx, state))
    if action == "row.delete":
        return with_workbook(state, flags, cwd, positionals, lambda wb, path, pos, fx: xl_row_delete(wb, path, pos, fx, state))
    if action == "col.add":
        return with_workbook(state, flags, cwd, positionals, lambda wb, path, pos, fx: xl_col_add(wb, path, pos, fx, state))
    if action == "col.delete":
        return with_workbook(state, flags, cwd, positionals, lambda wb, path, pos, fx: xl_col_delete(wb, path, pos, fx, state))
    if action == "merge":
        return with_workbook(state, flags, cwd, positionals, lambda wb, path, pos, fx: xl_merge(wb, path, pos, fx, state))
    if action == "unmerge":
        return with_workbook(state, flags, cwd, positionals, lambda wb, path, pos, fx: xl_unmerge(wb, path, pos, fx, state))
    if action == "export":
        return with_workbook(state, flags, cwd, positionals, lambda wb, path, pos, fx: xl_export(wb, path, fx, state, cwd))

    return fail(f"xl {action}: not supported by {BACKEND_ID}. Supported: {', '.join(SUPPORTED_ACTIONS)}", 127)


def xl_open(positionals: list[str], flags: dict[str, Any], cwd: str, state: dict[str, Any]) -> dict[str, Any]:
    raw = first(positionals) or str(flags.get("path") or "")
    path = resolve_user_path(raw, cwd)
    if not path:
        return fail("xl open: missing <path>", 2)
    if Path(path).suffix.lower() == ".xls":
        return fail("xl open: .xls is not supported by openpyxl backend; use .xlsx/.xlsm", 1)
    wb = load_book(path)
    remember_open(state, path, wb.active.title)
    save_state(state)
    return ok_json({"ok": True, "opened": path, "sheets": len(wb.sheetnames), "active": wb.active.title})


def xl_new(positionals: list[str], flags: dict[str, Any], cwd: str, state: dict[str, Any]) -> dict[str, Any]:
    raw = first(positionals) or str(flags.get("path") or "")
    path = resolve_user_path(raw, cwd) if raw else str(default_new_path())
    force = truthy(flags.get("force"))
    if Path(path).exists() and not force:
        return fail(f"xl new: {path} already exists; pass --force=true to overwrite", 1)
    wb = Workbook()
    ensure_parent(path)
    wb.save(path)
    remember_open(state, path, wb.active.title)
    save_state(state)
    return ok_json({"ok": True, "created": path, "sheets": len(wb.sheetnames), "active": wb.active.title})


def xl_save(positionals: list[str], flags: dict[str, Any], cwd: str, state: dict[str, Any]) -> dict[str, Any]:
    src = active_path(state, flags, cwd)
    if not src:
        return fail("xl save: no active workbook", 1)
    raw_dest = str(flags.get("path") or first(positionals) or "")
    if not raw_dest:
        return ok_json({"ok": True, "path": src, "note": "workbook changes are saved after each write"})
    dest = resolve_user_path(raw_dest, cwd)
    if not dest:
        return fail("xl save: invalid --path", 2)
    if Path(dest).exists() and dest != src and not truthy(flags.get("force")):
        return fail(f"xl save: {dest} already exists; pass --force=true to overwrite", 1)
    wb = load_book(src)
    ensure_parent(dest)
    wb.save(dest)
    remember_open(state, dest, active_sheet_name(state, src) or wb.active.title)
    save_state(state)
    return ok_json({"ok": True, "path": dest})


def xl_close(flags: dict[str, Any], state: dict[str, Any]) -> dict[str, Any]:
    path = state.get("activePath")
    if not path:
        return fail("xl close: no active workbook", 1)
    paths = [p for p in state.get("openPaths", []) if p != path]
    state["openPaths"] = paths
    state["activePath"] = paths[0] if paths else None
    save_state(state)
    return ok_json({"ok": True, "closed": path, "saved": flags.get("save") != "false"})


def xl_books(state: dict[str, Any]) -> dict[str, Any]:
    paths = [str(p) for p in state.get("openPaths", []) if Path(str(p)).exists()]
    state["openPaths"] = paths
    if state.get("activePath") not in paths:
        state["activePath"] = paths[0] if paths else None
    save_state(state)
    lines = ["path\tsheets\tactive"]
    for path in paths:
        try:
            wb = load_book(path)
            active = "*" if path == state.get("activePath") else ""
            lines.append(f"{path}\t{len(wb.sheetnames)}\t{active}")
        except Exception as exc:
            lines.append(f"{path}\tERROR: {exc}\t")
    return ok("\n".join(lines))


def xl_info(wb: Any, path: str, flags: dict[str, Any], state: dict[str, Any]) -> dict[str, Any]:
    active = active_sheet_name(state, path) or wb.active.title
    sheets: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        bounds = actual_used_bounds(ws)
        if bounds:
            min_row, min_col, max_row, max_col = bounds
            used = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
            rows = max_row - min_row + 1
            cols = max_col - min_col + 1
        else:
            used = "A1"
            rows = 0
            cols = 0
        sheets.append({"name": ws.title, "rows": rows, "cols": cols, "used": used, "active": ws.title == active})
    return ok_json({"ok": True, "path": path, "sheets": sheets, "active": active})


def xl_sheets(wb: Any, path: str) -> dict[str, Any]:
    active = active_sheet_name(load_state(), path) or wb.active.title
    lines = ["name\trows\tcols\tactive"]
    for ws in wb.worksheets:
        bounds = actual_used_bounds(ws)
        rows = max(0, bounds[2] - bounds[0] + 1) if bounds else 0
        cols = max(0, bounds[3] - bounds[1] + 1) if bounds else 0
        lines.append(f"{ws.title}\t{rows}\t{cols}\t{'*' if ws.title == active else ''}")
    return ok("\n".join(lines))


def xl_sheet(wb: Any, path: str, positionals: list[str], state: dict[str, Any]) -> dict[str, Any]:
    name = first(positionals)
    if not name:
        return fail("xl sheet: missing <name>", 2)
    if name not in wb.sheetnames:
        return fail(f'xl sheet: not found "{name}"', 1)
    state.setdefault("activeSheetByPath", {})[path] = name
    wb.active = wb.sheetnames.index(name)
    wb.save(path)
    save_state(state)
    return ok_json({"ok": True, "active": name})


def xl_sheet_add(wb: Any, path: str, flags: dict[str, Any], state: dict[str, Any]) -> dict[str, Any]:
    name = str(flags.get("name") or "")
    if not name:
        return fail("xl sheet.add: missing --name", 2)
    if name in wb.sheetnames:
        return fail(f'xl sheet.add: "{name}" already exists', 1)
    wb.create_sheet(name)
    state.setdefault("activeSheetByPath", {})[path] = name
    wb.active = wb.sheetnames.index(name)
    wb.save(path)
    save_state(state)
    return ok_json({"ok": True, "added": name, "totalSheets": len(wb.sheetnames)})


def xl_sheet_rename(wb: Any, path: str, positionals: list[str], flags: dict[str, Any], state: dict[str, Any]) -> dict[str, Any]:
    old = first(positionals)
    new = str(flags.get("name") or "")
    if not old or not new:
        return fail("xl sheet.rename: need <oldName> --name", 2)
    if old not in wb.sheetnames:
        return fail(f'xl sheet.rename: not found "{old}"', 1)
    if new in wb.sheetnames:
        return fail(f'xl sheet.rename: "{new}" already exists', 1)
    wb[old].title = new
    if active_sheet_name(state, path) == old:
        state.setdefault("activeSheetByPath", {})[path] = new
    wb.save(path)
    save_state(state)
    return ok_json({"ok": True, "from": old, "to": new})


def xl_sheet_delete(wb: Any, path: str, positionals: list[str], state: dict[str, Any]) -> dict[str, Any]:
    name = first(positionals)
    if not name:
        return fail("xl sheet.delete: missing <name>", 2)
    if name not in wb.sheetnames:
        return fail(f'xl sheet.delete: not found "{name}"', 1)
    if len(wb.sheetnames) == 1:
        return fail("xl sheet.delete: cannot delete the only sheet", 1)
    ws = wb[name]
    wb.remove(ws)
    if active_sheet_name(state, path) == name:
        state.setdefault("activeSheetByPath", {})[path] = wb.active.title
    wb.save(path)
    save_state(state)
    return ok_json({"ok": True, "deleted": name, "totalSheets": len(wb.sheetnames)})


def xl_used(wb: Any, path: str, flags: dict[str, Any], state: dict[str, Any]) -> dict[str, Any]:
    ws = selected_sheet(wb, path, flags, state)
    bounds = actual_used_bounds(ws)
    if not bounds:
        return ok_json({"sheet": ws.title, "range": "A1", "rows": 0, "cols": 0})
    min_row, min_col, max_row, max_col = bounds
    addr = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    return ok_json({"sheet": ws.title, "range": addr, "rows": max_row - min_row + 1, "cols": max_col - min_col + 1})


def xl_get(wb: Any, path: str, positionals: list[str], flags: dict[str, Any], state: dict[str, Any]) -> dict[str, Any]:
    raw = first(positionals) or str(flags.get("cell") or "")
    if not raw:
        return fail("xl get: missing <cell>", 2)
    sheet_name, address = parse_ref(raw)
    ws = selected_sheet(wb, path, flags, state, sheet_name)
    cell = ws[address]
    value = cell.value
    formula = value if isinstance(value, str) and value.startswith("=") else ""
    fmt = str(flags.get("format") or "json").lower()
    if fmt == "raw":
        return ok(cell_text(value))
    if fmt == "tsv":
        return ok(f"{cell_text(value)}\t{formula}")
    return ok_json({"cell": raw, "sheet": ws.title, "value": value, "formula": formula})


def xl_range(wb: Any, path: str, positionals: list[str], flags: dict[str, Any], state: dict[str, Any]) -> dict[str, Any]:
    raw = first(positionals) or str(flags.get("range") or "")
    if not raw:
        return fail("xl range: missing <range>", 2)
    sheet_name, address = parse_ref(raw)
    ws = selected_sheet(wb, path, flags, state, sheet_name)
    min_col, min_row, max_col, max_row = safe_range_bounds(address, "range")
    lines: list[str] = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        lines.append("\t".join(cell_text(cell.value) for cell in row))
    return ok("\n".join(lines))


def xl_find(wb: Any, path: str, positionals: list[str], flags: dict[str, Any], state: dict[str, Any]) -> dict[str, Any]:
    needle = str(flags.get("what") or first(positionals) or "")
    if not needle:
        return fail("xl find: missing keyword", 2)
    in_ref = str(flags.get("in") or "")
    sheet_name, address = parse_ref(in_ref) if in_ref else (None, "")
    ws = selected_sheet(wb, path, flags, state, sheet_name)
    if address:
        min_col, min_row, max_col, max_row = safe_range_bounds(address, "find")
    else:
        bounds = actual_used_bounds(ws)
        if not bounds:
            return ok("address\tvalue")
        min_row, min_col, max_row, max_col = bounds
    lines = ["address\tvalue"]
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            text = cell_text(cell.value)
            if needle in text:
                lines.append(f"{cell.coordinate}\t{text}")
    return ok("\n".join(lines))


def xl_set(wb: Any, path: str, positionals: list[str], flags: dict[str, Any], state: dict[str, Any]) -> dict[str, Any]:
    raw = first(positionals) or str(flags.get("cell") or "")
    if not raw:
        return fail("xl set: missing <cell>", 2)
    sheet_name, address = parse_ref(raw)
    ws = selected_sheet(wb, path, flags, state, sheet_name)
    if flags.get("formula") is not None:
        value = str(flags["formula"])
        ws[address] = value if value.startswith("=") else "=" + value
    elif flags.get("value") is not None:
        ws[address] = coerce_value(flags["value"], str(flags.get("type") or "auto"))
    else:
        return fail("xl set: missing --value or --formula", 2)
    wb.save(path)
    return ok_json({"ok": True, "sheet": ws.title, "cell": address})


def xl_fill(wb: Any, path: str, positionals: list[str], flags: dict[str, Any], state: dict[str, Any]) -> dict[str, Any]:
    raw = first(positionals) or str(flags.get("range") or "")
    if not raw:
        return fail("xl fill: missing <range>", 2)
    if flags.get("values") is None:
        return fail("xl fill: missing --values (TSV)", 2)
    sheet_name, address = parse_ref(raw)
    ws = selected_sheet(wb, path, flags, state, sheet_name)
    min_col, min_row, _, _ = safe_range_bounds(address, "fill")
    text = str(flags["values"]).replace("\\n", "\n")
    written = 0
    for r_offset, row_text in enumerate(text.splitlines()):
        for c_offset, raw_value in enumerate(row_text.split("\t")):
            ws.cell(row=min_row + r_offset, column=min_col + c_offset).value = coerce_value(raw_value, "auto")
            written += 1
    wb.save(path)
    return ok_json({"ok": True, "sheet": ws.title, "written": written})


def xl_clear(wb: Any, path: str, positionals: list[str], flags: dict[str, Any], state: dict[str, Any]) -> dict[str, Any]:
    raw = first(positionals) or str(flags.get("range") or "")
    if not raw:
        return fail("xl clear: missing <range>", 2)
    sheet_name, address = parse_ref(raw)
    ws = selected_sheet(wb, path, flags, state, sheet_name)
    min_col, min_row, max_col, max_row = safe_range_bounds(address, "clear")
    cleared = 0
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                cleared += 1
            cell.value = None
    wb.save(path)
    return ok_json({"ok": True, "sheet": ws.title, "cleared": cleared})


def xl_fmt(wb: Any, path: str, positionals: list[str], flags: dict[str, Any], state: dict[str, Any]) -> dict[str, Any]:
    raw = first(positionals) or str(flags.get("range") or "")
    if not raw:
        return fail("xl fmt: missing <range>", 2)
    sheet_name, address = parse_ref(raw)
    ws = selected_sheet(wb, path, flags, state, sheet_name)
    applied: dict[str, Any] = {}
    for cell in cells_in_range(ws, address, "fmt"):
        font_kwargs: dict[str, Any] = {
            "name": cell.font.name,
            "size": cell.font.sz,
            "bold": cell.font.bold,
            "italic": cell.font.italic,
            "color": cell.font.color,
        }
        if "bold" in flags:
            font_kwargs["bold"] = truthy(flags.get("bold"))
            applied["bold"] = font_kwargs["bold"]
        if "italic" in flags:
            font_kwargs["italic"] = truthy(flags.get("italic"))
            applied["italic"] = font_kwargs["italic"]
        if flags.get("size") is not None:
            font_kwargs["size"] = float(str(flags["size"]))
            applied["size"] = font_kwargs["size"]
        if flags.get("color") is not None:
            font_kwargs["color"] = normalize_color(str(flags["color"]))
            applied["color"] = str(flags["color"])
        cell.font = Font(**font_kwargs)
        if flags.get("bg") is not None:
            cell.fill = PatternFill(fill_type="solid", fgColor=normalize_color(str(flags["bg"])))
            applied["bg"] = str(flags["bg"])
        if flags.get("align") is not None:
            cell.alignment = Alignment(horizontal=str(flags["align"]))
            applied["align"] = str(flags["align"])
        if flags.get("numfmt") is not None:
            cell.number_format = str(flags["numfmt"])
            applied["numfmt"] = str(flags["numfmt"])
        if flags.get("border") is not None:
            side = Side(style="thin", color="FF000000")
            cell.border = Border(left=side, right=side, top=side, bottom=side)
            applied["border"] = str(flags["border"])
    wb.save(path)
    return ok_json({"ok": True, "sheet": ws.title, "range": address, "applied": applied})


def xl_width(wb: Any, path: str, positionals: list[str], flags: dict[str, Any], state: dict[str, Any]) -> dict[str, Any]:
    col = first(positionals)
    size = flags.get("size", flags.get("width"))
    if not col or size is None:
        return fail("xl width: need <column> --size", 2)
    ws = selected_sheet(wb, path, flags, state)
    ws.column_dimensions[col.upper()].width = float(str(size))
    wb.save(path)
    return ok_json({"ok": True, "sheet": ws.title, "column": col.upper(), "size": float(str(size))})


def xl_row_add(wb: Any, path: str, positionals: list[str], flags: dict[str, Any], state: dict[str, Any]) -> dict[str, Any]:
    row = positive_int(first(positionals), "row.add <row>")
    count = positive_int(flags.get("count") or "1", "--count")
    ws = selected_sheet(wb, path, flags, state)
    ws.insert_rows(row, count)
    wb.save(path)
    return ok_json({"ok": True, "sheet": ws.title, "inserted": count, "at": row})


def xl_row_delete(wb: Any, path: str, positionals: list[str], flags: dict[str, Any], state: dict[str, Any]) -> dict[str, Any]:
    row = positive_int(first(positionals), "row.delete <row>")
    count = positive_int(flags.get("count") or "1", "--count")
    ws = selected_sheet(wb, path, flags, state)
    ws.delete_rows(row, count)
    wb.save(path)
    return ok_json({"ok": True, "sheet": ws.title, "deleted": count, "at": row})


def xl_col_add(wb: Any, path: str, positionals: list[str], flags: dict[str, Any], state: dict[str, Any]) -> dict[str, Any]:
    col = first(positionals)
    if not col:
        return fail("xl col.add: missing <column>", 2)
    count = positive_int(flags.get("count") or "1", "--count")
    ws = selected_sheet(wb, path, flags, state)
    ws.insert_cols(range_boundaries(f"{col}1:{col}1")[0], count)
    wb.save(path)
    return ok_json({"ok": True, "sheet": ws.title, "inserted": count, "at": col.upper()})


def xl_col_delete(wb: Any, path: str, positionals: list[str], flags: dict[str, Any], state: dict[str, Any]) -> dict[str, Any]:
    col = first(positionals)
    if not col:
        return fail("xl col.delete: missing <column>", 2)
    count = positive_int(flags.get("count") or "1", "--count")
    ws = selected_sheet(wb, path, flags, state)
    ws.delete_cols(range_boundaries(f"{col}1:{col}1")[0], count)
    wb.save(path)
    return ok_json({"ok": True, "sheet": ws.title, "deleted": count, "at": col.upper()})


def xl_merge(wb: Any, path: str, positionals: list[str], flags: dict[str, Any], state: dict[str, Any]) -> dict[str, Any]:
    raw = first(positionals) or str(flags.get("range") or "")
    if not raw:
        return fail("xl merge: missing <range>", 2)
    sheet_name, address = parse_ref(raw)
    ws = selected_sheet(wb, path, flags, state, sheet_name)
    ws.merge_cells(address)
    wb.save(path)
    return ok_json({"ok": True, "sheet": ws.title, "merged": address})


def xl_unmerge(wb: Any, path: str, positionals: list[str], flags: dict[str, Any], state: dict[str, Any]) -> dict[str, Any]:
    raw = first(positionals) or str(flags.get("range") or "")
    if not raw:
        return fail("xl unmerge: missing <range>", 2)
    sheet_name, address = parse_ref(raw)
    ws = selected_sheet(wb, path, flags, state, sheet_name)
    ws.unmerge_cells(address)
    wb.save(path)
    return ok_json({"ok": True, "sheet": ws.title, "unmerged": address})


def xl_export(wb: Any, path: str, flags: dict[str, Any], state: dict[str, Any], cwd: str) -> dict[str, Any]:
    fmt = str(flags.get("format") or "").lower()
    raw_out = str(flags.get("path") or "")
    if not fmt or not raw_out:
        return fail("xl export: need --format and --path", 2)
    if fmt != "csv":
        return fail('xl export: python-openpyxl backend supports only --format=csv on headless Linux', 1)
    out_path = resolve_user_path(raw_out, cwd)
    if Path(out_path).exists() and not truthy(flags.get("force")):
        return fail(f"xl export: {out_path} already exists; pass --force=true to overwrite", 1)
    ws = selected_sheet(wb, path, flags, state)
    bounds = actual_used_bounds(ws)
    ensure_parent(out_path)
    with open(out_path, "w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        if bounds:
            min_row, min_col, max_row, max_col = bounds
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                writer.writerow([cell.value for cell in row])
    return ok_json({"ok": True, "exported": out_path, "format": "csv", "sheet": ws.title})


def with_workbook(
    state: dict[str, Any],
    flags: dict[str, Any],
    cwd: str,
    positionals: list[str],
    fn: Any,
) -> dict[str, Any]:
    effective_positionals, effective_flags = split_workbook_positionals(positionals, flags)
    path = active_path(state, effective_flags, cwd)
    if not path:
        return fail("xl: no active workbook; run `xl open <file.xlsx>` first", 1)
    wb = load_book(path)
    return fn(wb, path, effective_positionals, effective_flags)


def split_workbook_positionals(
    positionals: list[str],
    flags: dict[str, Any],
) -> tuple[list[str], dict[str, Any]]:
    effective_flags = dict(flags)
    effective_positionals = list(positionals)
    if not (effective_flags.get("book") or effective_flags.get("workbook")) and effective_positionals:
        first_arg = effective_positionals[0]
        if looks_like_workbook_path(first_arg):
            effective_flags["book"] = first_arg
            effective_positionals = effective_positionals[1:]
    return effective_positionals, effective_flags


def looks_like_workbook_path(raw: str) -> bool:
    suffix = Path(os.path.expanduser(raw)).suffix.lower()
    return suffix in (".xlsx", ".xlsm", ".xltx", ".xltm", ".xls")


def load_book(path: str) -> Any:
    if not Path(path).exists():
        raise FileNotFoundError(path)
    suffix = Path(path).suffix.lower()
    if suffix == ".xls":
        raise RuntimeError(".xls is not supported by openpyxl backend; use .xlsx/.xlsm")
    return load_workbook(path, keep_vba=suffix == ".xlsm")


def selected_sheet(
    wb: Any,
    path: str,
    flags: dict[str, Any],
    state: dict[str, Any],
    explicit: str | None = None,
) -> Any:
    name = explicit or str(flags.get("sheet") or "") or active_sheet_name(state, path)
    if name:
        if name not in wb.sheetnames:
            raise RuntimeError(f'sheet "{name}" not found')
        return wb[name]
    return wb.active


def parse_ref(raw: str) -> tuple[str | None, str]:
    value = raw.strip()
    if "!" in value:
        left, right = value.rsplit("!", 1)
        return left.strip("'"), right
    if "@" in value:
        left, right = value.rsplit("@", 1)
        if left and right:
            return left, right
    return None, value


def safe_range_bounds(address: str, action: str) -> tuple[int, int, int, int]:
    try:
        return range_boundaries(address)
    except Exception as exc:
        raise RuntimeError(f"xl {action}: invalid range {address}: {exc}") from exc


def cells_in_range(ws: Any, address: str, action: str) -> Iterable[Any]:
    min_col, min_row, max_col, max_row = safe_range_bounds(address, action)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        yield from row


def actual_used_bounds(ws: Any) -> tuple[int, int, int, int] | None:
    min_row = min_col = None
    max_row = max_col = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            min_row = cell.row if min_row is None else min(min_row, cell.row)
            min_col = cell.column if min_col is None else min(min_col, cell.column)
            max_row = max(max_row, cell.row)
            max_col = max(max_col, cell.column)
    if min_row is None or min_col is None:
        return None
    return min_row, min_col, max_row, max_col


def load_state() -> dict[str, Any]:
    try:
        data = json.loads(STATE_FILE.read_text(encoding="utf-8"))
        if isinstance(data, dict):
            data.setdefault("openPaths", [])
            data.setdefault("activeSheetByPath", {})
            return data
    except Exception:
        pass
    return {"activePath": None, "openPaths": [], "activeSheetByPath": {}}


def save_state(state: dict[str, Any]) -> None:
    STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
    STATE_FILE.write_text(json.dumps(state, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def remember_open(state: dict[str, Any], path: str, sheet: str | None) -> None:
    paths = [str(p) for p in state.get("openPaths", []) if str(p) != path]
    paths.append(path)
    state["openPaths"] = paths
    state["activePath"] = path
    if sheet:
        state.setdefault("activeSheetByPath", {})[path] = sheet


def active_path(state: dict[str, Any], flags: dict[str, Any], cwd: str) -> str | None:
    override = flags.get("book") or flags.get("workbook")
    if override:
        return resolve_user_path(str(override), cwd)
    path = state.get("activePath")
    return str(path) if path else None


def active_sheet_name(state: dict[str, Any], path: str) -> str | None:
    by_path = state.get("activeSheetByPath") or {}
    value = by_path.get(path) if isinstance(by_path, dict) else None
    return str(value) if value else None


def resolve_user_path(raw: str, cwd: str) -> str:
    if not raw:
        return ""
    expanded = Path(os.path.expanduser(raw))
    if not expanded.is_absolute():
        expanded = Path(cwd) / expanded
    return str(expanded.resolve(strict=False))


def default_new_path() -> Path:
    STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
    return STATE_FILE.parent / f"Untitled-{int(time.time())}.xlsx"


def ensure_parent(path: str) -> None:
    Path(path).parent.mkdir(parents=True, exist_ok=True)


def coerce_value(value: Any, kind: str) -> Any:
    if kind == "text":
        return str(value)
    if kind == "number":
        try:
            return float(str(value))
        except ValueError as exc:
            raise RuntimeError(f'xl set: --type=number but "{value}" is not numeric') from exc
    if isinstance(value, (int, float, bool)):
        return value
    text = str(value)
    if text == "":
        return ""
    try:
        num = float(text)
        return int(num) if num.is_integer() else num
    except ValueError:
        return text


def normalize_color(raw: str) -> str:
    s = raw.strip().lower()
    if s in COLOR_NAMES:
        return COLOR_NAMES[s]
    if s.startswith("#") and len(s) == 7:
        return "FF" + s[1:].upper()
    if "," in s:
        parts = [int(p.strip()) for p in s.split(",")]
        if len(parts) == 3 and all(0 <= p <= 255 for p in parts):
            return "FF" + "".join(f"{p:02X}" for p in parts)
    if len(s) == 6 and all(ch in "0123456789abcdef" for ch in s):
        return "FF" + s.upper()
    raise RuntimeError(f'xl fmt: invalid color "{raw}"')


def positive_int(value: Any, label: str) -> int:
    try:
        n = int(str(value))
    except Exception as exc:
        raise RuntimeError(f"xl {label}: expected positive integer") from exc
    if n < 1:
        raise RuntimeError(f"xl {label}: expected positive integer")
    return n


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def first(values: list[str]) -> str:
    return values[0] if values else ""


def truthy(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).lower() not in ("", "0", "false", "no", "off")


def ok(stdout: str) -> dict[str, Any]:
    return {"ok": True, "stdout": stdout, "stderr": "", "exit_code": 0}


def ok_json(obj: dict[str, Any]) -> dict[str, Any]:
    return ok(json.dumps(obj, ensure_ascii=False, indent=2))


def fail(message: str, code: int) -> dict[str, Any]:
    return {"ok": False, "stdout": "", "stderr": message, "exit_code": code}


if __name__ == "__main__":
    raise SystemExit(main())
